"""for generating ticket request forms"""
import os
import xlrd
import datetime
import mailmerge
import openpyxl
import math

today = datetime.date.today().strftime("%m/%d/%Y")

downloadFolder = 'C:\\Users\\DanFang\\Downloads'
os.chdir(downloadFolder)

dated_files = [(os.path.getmtime(fn), os.path.basename(fn))
               for fn in os.listdir(downloadFolder) if fn.lower().endswith('.xlsx')]
dated_files.sort()
dated_files.reverse()
excel = dated_files[0][1]

open_excel = xlrd.open_workbook(excel)
sheet = open_excel.sheet_by_index(0)
keys = [sheet.cell(0, col_index).value for col_index in range(sheet.ncols)]

ticket_requests = []
for row_index in range(1, sheet.nrows):
    d = {keys[col_index]: sheet.cell(row_index, col_index).value for col_index in range(sheet.ncols)}
    ticket_requests.append(d)

#print(ticket_requests)

class Request(object):
    def __init__(self, event = "", agency = "", preferred_date_total = "", preferred_date = "", preferred_time = "", alternate_date = "", alternate_time = "", tickets_requested = "", number_children = "", number_adults = "", chaperone_name = "", chaperone_cell = "", notes = "", kids_youngest = "", kids_oldest = "", address = "", city = "", state = "", postal = "", email = "", phone = "", fax = "", primary_contact = "", transportation_type = ""):
        self.event = event
        self.agency = agency
        self.preferred_date_total = preferred_date_total
        self.preferred_date = preferred_date
        self.preferred_time = preferred_time
        self.alternate_date = alternate_date
        self.alternate_time = alternate_time
        self.tickets_requested = tickets_requested
        self.number_children = number_children
        self.number_adults = number_adults
        self.chaperone_name = chaperone_name
        self.chaperone_cell = chaperone_cell
        self.notes = notes
        self.kids_youngest = kids_youngest
        self.kids_oldest = kids_oldest
        self.address = address
        self.city = city
        self.state = state
        self.postal = postal
        self.email = email
        self.phone = phone
        self.fax = fax
        self.primary_contact = primary_contact
        self.transportation_type = transportation_type


#print(ticket_requests[0])

#request_objects = []
#for request in ticket_requests:
    #request_objects.append(Request(request["Event"]))

#print(len(request_objects))

request_objects = []

for request in ticket_requests:
    pref_date_string = xlrd.xldate.xldate_as_datetime(request['Preferred Date'], open_excel.datemode)
    request["Pref Date"] = datetime.datetime.date(pref_date_string)
    request["Pref Time"] = datetime.datetime.time(pref_date_string)
    request["Alt Date"] = ""
    request["Alt Time"] = ""
    if request['Alternate Date'] != '':
        alt_date_string = xlrd.xldate.xldate_as_datetime(request['Alternate Date'], open_excel.datemode)
        request["Alt Date"] = datetime.datetime.date(alt_date_string)
        request["Alt Time"] = datetime.datetime.time(alt_date_string)

    primary_contact = " ".join(request["Primary Contact (Agency) (Account)"].split(", ")[::-1])
    request["Primary Contact (Agency) (Account)"] = primary_contact
    #print(request["Primary Contact (Agency) (Account)"])

    request_objects.append(Request(request["Event"], request["Agency"], pref_date_string, request["Pref Date"], request["Pref Time"], request["Alt Date"], request["Alt Time"], request["Tickets Requested"], request["Number of Children"], request["Total Adults"], request["Chaperone Name"], request["Chaperone Cell Phone"], request["Notes"], request["Kids Youngest Age"], request["Kids Oldest Age"], request["Address Line 1 (Agency) (Account)"], request["City (Agency) (Account)"], request["State or Province (Agency) (Account)"], request["Postal Code (Agency) (Account)"], request["Email (Agency) (Account)"], request["Main Phone (Agency) (Account)"], request["Fax (Agency) (Account)"], request["Primary Contact (Agency) (Account)"], request["Transportation Type"]))

out_folder = "C:\\Users\\DanFang\\Desktop\\Requests"
os.chdir(out_folder)

#Region Carnegie Science Center

csc_requests = []
for request in request_objects:
    if "Carnegie Science Center" in request.event:
        csc_requests.append(request)

print('CSC:', len(csc_requests))

csc_requests.sort(key=lambda r: r.preferred_date)

if len(csc_requests) > 0:
    cscdoc = mailmerge.MailMerge('csctemplate.docx')
    #print(cscdoc.get_merge_fields())

    csc_merge = []
    for request in csc_requests:
        departure_string = request.preferred_date_total + datetime.timedelta(hours=4)
        departure_time = datetime.datetime.time(departure_string)

        if request.alternate_date != "":
            csc_merge.append({'PreferredTime': request.preferred_time.strftime("%#I:%M %p"),
                              'DepartureTime': departure_time.strftime("%#I:%M %p"),
                              'DateToday': today,
                              'PreferredDate': request.preferred_date.strftime("%#m/%#d/%Y"),
                              'Notes': request.notes,
                              'MainContact': request.primary_contact,
                              'Agency': request.agency,
                              'city': request.city,
                              'Chaperone': request.chaperone_name,
                              'Zip': request.postal,
                              'Alternate': request.alternate_date.strftime("%#m/%#d/%Y") + " " + request.alternate_time.strftime("%#I:%M %p"),
                              'NumberAdults': str(int(request.number_adults)),
                              'Number': str(int(request.number_children)),
                              'State': request.state,
                              'ChaperoneCell': request.chaperone_cell,
                              'Address': request.address,
                              'Total': str(int(request.tickets_requested)),
                              'TransportationType': request.transportation_type,
                              'AgencyPhone': request.phone})
        else:
            csc_merge.append({'PreferredTime': request.preferred_time.strftime("%#I:%M %p"),
                              'DepartureTime': departure_time.strftime("%#I:%M %p"),
                              'DateToday': today,
                              'PreferredDate': request.preferred_date.strftime("%#m/%#d/%Y"),
                              'Notes': request.notes,
                              'MainContact': request.primary_contact, 'Agency': request.agency,
                              'city': request.city,
                              'Chaperone': request.chaperone_name,
                              'Zip': request.postal,
                              'Alternate': "",
                              'NumberAdults': str(int(request.number_adults)),
                              'Number': str(int(request.number_children)),
                              'State': request.state,
                              'ChaperoneCell': request.chaperone_cell,
                              'Address': request.address,
                              'Total': str(int(request.tickets_requested)),
                              'TransportationType': request.transportation_type,
                              'AgencyPhone': request.phone})

    cscdoc.merge_pages(csc_merge)

    csc_request_title = datetime.date.today().strftime("%Y-%m-%d") + ' TFK CSC Ticket Requests.docx'

    cscdoc.write(csc_request_title)

#Region Children's Museum

cmpgh_requests = []
for request in request_objects:
    if "Children's Museum" in request.event:
        cmpgh_requests.append(request)

print('Childrens Museum:', len(cmpgh_requests))

cmpgh_requests.sort(key=lambda r: r.preferred_date)

if len(cmpgh_requests) > 0:
    cmpghdoc = mailmerge.MailMerge('childrensmuseumtemplate.docx')
    #print(cmpghdoc.get_merge_fields())

    cmpgh_merge = []
    for request in cmpgh_requests:
        departure_string = request.preferred_date_total + datetime.timedelta(hours=4)
        departure_time = datetime.datetime.time(departure_string)
        cmpgh_merge.append({'PreferredTime': request.preferred_time.strftime("%#I:%M %p"),
                            'DepartureTime': departure_time.strftime("%#I:%M %p"),
                            'PreferredDate': request.preferred_date.strftime("%#m/%#d/%Y"),
                            'Notes': request.notes,
                            'Agency': request.agency,
                            'City': request.city,
                            'Chaperone': request.chaperone_name,
                            'Zip': request.postal,
                            'NumberAdults': str(int(request.number_adults)),
                            'NumberChildren': str(int(request.number_children)),
                            'State': request.state,
                            'ChaperonePhone': request.chaperone_cell,
                            'Address': request.address,
                            'Transportation': request.transportation_type,
                            'KidsYoungest': str(int(request.kids_youngest)),
                            'KidsOldest': str(int(request.kids_oldest)) })

    cmpghdoc.merge_pages(cmpgh_merge)

    cmpgh_request_title = datetime.date.today().strftime("%Y-%m-%d") + ' TFK Children\'s Museum Ticket Requests.docx'

    cmpghdoc.write(cmpgh_request_title)

#Region History Center

heinz_requests = []
for request in request_objects:
    if "Heinz History Center" in request.event:
        heinz_requests.append(request)

print('Heinz History:', len(heinz_requests))

heinz_requests.sort(key=lambda r: r.preferred_date)

if len(heinz_requests) > 0:
    heinzdoc = mailmerge.MailMerge('heinztemplate.docx')
    # print(heinzdoc.get_merge_fields())

    heinz_merge = []
    for request in heinz_requests:
        departure_string = request.preferred_date_total + datetime.timedelta(hours=2)
        departure_time = datetime.datetime.time(departure_string)
        if request.alternate_date != "":
            heinz_merge.append({'PreferredTime': request.preferred_time.strftime("%#I:%M %p"),
                                'DepartureTime': departure_time.strftime("%#I:%M %p"),
                                'PreferredDate': request.preferred_date.strftime("%#m/%#d/%Y"),
                                'AlternateDate': request.alternate_date.strftime("%#m/%#d/%Y") + " " + request.alternate_time.strftime("%#I:%M %p"),
                                'Notes': request.notes,
                                'Agency': request.agency,
                                'Chaperone': request.chaperone_name,
                                'NumberAdults': str(int(request.number_adults)),
                                'NumberChildren': str(int(request.number_children)),
                                'ChaperonePhone': request.chaperone_cell,
                                'KidsYoungest': str(int(request.kids_youngest)),
                                'KidsOldest': str(int(request.kids_oldest)),
                                'Fax': request.fax})
        else:
            heinz_merge.append({'PreferredTime': request.preferred_time.strftime("%#I:%M %p"),
                                'DepartureTime': departure_time.strftime("%#I:%M %p"),
                                'PreferredDate': request.preferred_date.strftime("%#m/%#d/%Y"),
                                'AlternateDate': "",
                                'Notes': request.notes,
                                'Agency': request.agency,
                                'Chaperone': request.chaperone_name,
                                'NumberAdults': str(int(request.number_adults)),
                                'NumberChildren': str(int(request.number_children)),
                                'ChaperonePhone': request.chaperone_cell,
                                'KidsYoungest': str(int(request.kids_youngest)),
                                'KidsOldest': str(int(request.kids_oldest)),
                                'Fax': request.fax})

    heinzdoc.merge_pages(heinz_merge)

    heinz_request_title = datetime.date.today().strftime("%Y-%m-%d") + ' TFK Heinz History Ticket Requests.docx'

    heinzdoc.write(heinz_request_title)

# Region Mattress Factory

mattress_requests = []
for request in request_objects:
    if "Mattress Factory" in request.event:
        mattress_requests.append(request)

print('Mattress Factory:', len(mattress_requests))

mattress_requests.sort(key=lambda r: r.preferred_date)

if len(mattress_requests) > 0:
    mattressdoc = mailmerge.MailMerge('mattresstemplate.docx')
    # print(mattressdoc.get_merge_fields())

    mattress_merge = []
    for request in mattress_requests:
        mattress_merge.append({'PreferredDateTime': request.preferred_date.strftime("%#m/%#d/%Y") + " " + request.preferred_time.strftime("%#I:%M %p"),
                               'NumberAdults': str(int(request.number_adults)),
                               'Agency': request.agency,
                               'Chaperone': request.chaperone_name,
                               'NumberChildren': str(int(request.number_children)),
                               'ChaperonePhone': request.chaperone_cell,
                               "Main": request.primary_contact,
                               "Email": request.email,
                               "Notes": request.notes,
                               "TotalTickets": str(int(request.tickets_requested))})

    mattressdoc.merge_pages(mattress_merge)

    mattress_request_title = datetime.date.today().strftime("%Y-%m-%d") + ' TFK Mattress Factory Ticket Requests.docx'

    mattressdoc.write(mattress_request_title)


# Region Carnegie Museums

carnegie_requests = []
for request in request_objects:
    if "Carnegie Museums" in request.event:
        carnegie_requests.append(request)

print('Carnegie Museums:', len(carnegie_requests))

carnegie_requests.sort(key=lambda r: r.preferred_date)

if len(carnegie_requests) > 0:
    carnegiedoc = mailmerge.MailMerge('carnegiemuseumstemplate.docx')
    #print(cscdoc.get_merge_fields())

    carnegie_merge = []
    for request in carnegie_requests:

        carnegie_merge.append({'PreferredTime': request.preferred_time.strftime("%#I:%M %p"),
                               'PreferredDate': request.preferred_date.strftime("%#m/%#d/%Y"),
                               'Notes': request.notes,
                               'Agency': request.agency,
                               'City': request.city,
                               'ChaperoneName': request.chaperone_name,
                               'Zip': request.postal,
                               'NumAdults': str(int(request.number_adults)),
                               'NumChildren': str(int(request.number_children)),
                               'State': request.state,
                               'ChaperoneCell': request.chaperone_cell,
                               'Address': request.address,
                               'Youngest': str(int(request.kids_youngest)),
                               'Transportation': request.transportation_type,
                               'Oldest': str(int(request.kids_oldest))})


    carnegiedoc.merge_pages(carnegie_merge)

    carnegie_request_title = datetime.date.today().strftime("%Y-%m-%d") + ' TFK Carnegie Museums Requests.docx'

    carnegiedoc.write(carnegie_request_title)


# Region Aviary

aviary_requests = []
for request in request_objects:
    if "National Aviary" in request.event:
        aviary_requests.append(request)

print('Aviary:', len(aviary_requests))

aviary_requests.sort(key=lambda r: r.preferred_date)

if len(aviary_requests) > 0:

    large_font = openpyxl.styles.Font(name='Calibri',size=11)
    large_bold = openpyxl.styles.Font(name='Calibri',size=11, bold=True)
    small_font = openpyxl.styles.Font(name='Calibri',size=9)
    small_bold = openpyxl.styles.Font(name='Calibri',size=9, bold=True)
    fill = openpyxl.styles.PatternFill(fill_type='solid', start_color='00C5D9F1', end_color='00C5D9F1')
    border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),right=openpyxl.styles.Side(style='thin'),top=openpyxl.styles.Side(style='thin'),bottom=openpyxl.styles.Side(style='thin'))

    out_workbook = openpyxl.Workbook()
    sheet = out_workbook.active

    sheet["E1"] = "Scheduled/Reserved"
    sheet["E1"].font = large_font
    sheet["E1"].fill = fill
    sheet["E1"].border = border
    sheet.merge_cells('E1:H1')
    sheet["I1"] = "Actual Counts"
    sheet["I1"].font = large_bold
    sheet["I1"].fill = fill
    sheet["I1"].border = border
    sheet.merge_cells('I1:K1')
    sheet["L1"] = "MONTHLY"
    sheet["L1"].font = large_font
    sheet["L1"].fill = fill
    sheet["L1"].border = border

    sheet['A2'] = "DATE"
    sheet["B2"] = "TIME"
    sheet["C2"] = "ORG"
    sheet["D2"] = "AGE RANGE"
    sheet["E2"] = "KIDS"
    sheet["F2"] = "ADULTS"
    sheet["G2"] = "Total"
    sheet["H2"] = "MONTHLY"
    sheet["I2"] = "KIDS"
    sheet["J2"] = "ADULTS"
    sheet["K2"] = "Total"
    sheet["L2"] = "TOTAL"
    sheet["M2"] = "CHAPERONE NAME"
    sheet["N2"] = "CHAPERONE PHONE"
    sheet["O2"] = "ALTERNATE DATE"
    sheet["P2"] = "NOTES"

    column = 1

    for i in range(1, 17):
        sheet.cell(row=2, column=column).font = small_bold
        sheet.cell(row=2, column=column).fill = fill
        sheet.cell(row=2, column=column).border = border
        column += 1

    row_count = 3

    for request in aviary_requests:
        sheet['A' + str(row_count)] = request.preferred_date.strftime("%#m/%#d/%Y")
        sheet["B" + str(row_count)] = request.preferred_time.strftime("%#I:%M %p")
        sheet["C" + str(row_count)] = request.agency
        sheet["D" + str(row_count)] = str(int(request.kids_youngest)) + " to " + str(int(request.kids_oldest))
        sheet["E" + str(row_count)] = request.number_children
        sheet["F" + str(row_count)] = request.number_adults
        sheet["G" + str(row_count)] = request.tickets_requested
        # sheet["H" + str(row_count)] = "MONTHLY"
        # sheet["I" + str(row_count)] = "KIDS"
        # sheet["J" + str(row_count)] = "ADULTS"
        # sheet["K" + str(row_count)] = "Total"
        # sheet["L" + str(row_count)] = "TOTAL"
        sheet["M" + str(row_count)] = request.chaperone_name
        sheet["N" + str(row_count)] = request.chaperone_cell
        if request.alternate_date != "":
            sheet["O" + str(row_count)] = request.alternate_date.strftime("%#m/%#d/%Y") + " " + request.alternate_time.strftime("%#I:%M %p")
        sheet["P" + str(row_count)] = request.notes

        column = 1

        for i in range(1, 17):
            sheet.cell(row=row_count, column=column).font = small_font
            sheet.cell(row=row_count, column=column).border = border
            column += 1

        row_count += 1

    sheet["E" + str(row_count)] = "Kids"
    sheet["F" + str(row_count)] = "Adults"
    sheet["G" + str(row_count)] = "Total"
    # sheet["H" + str(row_count)] = 0
    sheet["I" + str(row_count)] = "Kids"
    sheet["J" + str(row_count)] = "Adults"
    sheet["K" + str(row_count)] = "Total"
    # sheet["L" + str(row_count)] = 0

    sheet["E" + str(row_count)].font = small_bold
    sheet["F" + str(row_count)].font = small_bold
    sheet["G" + str(row_count)].font = small_bold
    # sheet["H" + str(row_count)] = 0
    sheet["I" + str(row_count)].font = small_bold
    sheet["J" + str(row_count)].font = small_bold
    sheet["K" + str(row_count)].font = small_bold
    # sheet["L" + str(row_count)] = 0

    column = 5
    for i in range(5, 13):
        sheet.cell(row=row_count, column=column).font = small_bold
        sheet.cell(row=row_count, column=column).border = border
        column += 1

    row_count += 1

    sheet["E" + str(row_count)] = sum(x.number_children for x in aviary_requests)
    sheet["F" + str(row_count)] = sum(x.number_adults for x in aviary_requests)
    sheet["G" + str(row_count)] = sum(x.tickets_requested for x in aviary_requests)
    sheet["H" + str(row_count)] = 0
    sheet["I" + str(row_count)] = 0
    sheet["J" + str(row_count)] = 0
    sheet["K" + str(row_count)] = 0
    sheet["L" + str(row_count)] = 0

    column = 5
    for i in range(5, 13):
        sheet.cell(row=row_count, column=column).font = small_font
        sheet.cell(row=row_count, column=column).border = border
        column += 1

    for i in range(1, row_count + 1):
        sheet.row_dimensions[i].height = 12

    out_workbook.save(datetime.date.today().strftime("%Y-%m-%d") + ' TFK National Aviary Requests.xlsx')

# region PNC

pnc_requests = []
for request in request_objects:
    if "PNC Park Tours" in request.event:
        pnc_requests.append(request)

print('PNC:', len(pnc_requests))

pnc_requests.sort(key=lambda r: r.preferred_date)

if len(pnc_requests) > 0:
    pncdoc = mailmerge.MailMerge('pnctemplate.docx')

    pnc_merge = []
    for request in pnc_requests:

        numChildrenValue = request.number_children * 7

        numChaperones = int(request.number_adults)

        if (numChaperones > math.ceil(request.number_children / 12)):
            numChaperones = int(math.ceil(request.number_children / 12))

        numAdults = 0
        if request.number_adults - numChaperones > 0:
            numAdults = int(request.number_adults - numChaperones)

        numAdultValue =numAdults * 7

        if request.alternate_date != "":
            pnc_merge.append({'preferredDateandTime': request.preferred_date.strftime("%#m/%#d/%Y") + " " + request.preferred_time.strftime("%#I:%M %p") + "/" + request.alternate_date.strftime("%#m/%#d/%Y") + " " + request.alternate_time.strftime("%#I:%M %p"),
                              'notes': request.notes,
                              'mainContact': request.primary_contact,
                              'agencyName': request.agency,
                              'agencyCity': request.city,
                              'chaperone': request.chaperone_name,
                              'agencyZip': request.postal,
                              'agencyState': request.state,
                              'chaperonePhone': request.chaperone_cell,
                              'agencyAddress': request.address,
                              'agencyPhone': request.phone,
                              'agencyFax': request.fax,
                              'agencyEmail': request.email,
                              'youngest': str(int(request.kids_youngest)),
                              'oldest': str(int(request.kids_oldest)),
                              'numChildren': str(int(request.number_children)),
                              'numChildrenValue': '{0:.2f}'.format(numChildrenValue),
                              'numChaperone': str(numChaperones),
                              'numAdults': str(numAdults),
                              'numAdultValue': '{0:.2f}'.format(numAdultValue),
                              'numTotal': str(int(request.tickets_requested)),
                              'numTotalValue': '{0:.2f}'.format(numChildrenValue + numAdultValue)
                              })
        else:
            pnc_merge.append({'preferredDateandTime': request.preferred_date.strftime("%#m/%#d/%Y") + " " + request.preferred_time.strftime("%#I:%M %p"),
                              'notes': request.notes,
                              'mainContact': request.primary_contact,
                              'agencyName': request.agency,
                              'agencyCity': request.city,
                              'chaperone': request.chaperone_name,
                              'agencyZip': request.postal,
                              'agencyState': request.state,
                              'chaperonePhone': request.chaperone_cell,
                              'agencyAddress': request.address,
                              'agencyPhone': request.phone,
                              'agencyFax': request.fax,
                              'agencyEmail': request.email,
                              'youngest': str(int(request.kids_youngest)),
                              'oldest': str(int(request.kids_oldest)),
                              'numChildren': str(int(request.number_children)),
                              'numChildrenValue': '{0:.2f}'.format(numChildrenValue),
                              'numChaperone': str(numChaperones),
                              'numAdults': str(numAdults),
                              'numAdultValue': '{0:.2f}'.format(numAdultValue),
                              'numTotal': str(int(request.tickets_requested)),
                              'numTotalValue': '{0:.2f}'.format(numChildrenValue + numAdultValue)
                              })

    pncdoc.merge_pages(pnc_merge)

    pnc_request_title = datetime.date.today().strftime("%Y-%m-%d") + ' TFK PNC Park Tours Ticket Requests.docx'

    pncdoc.write(pnc_request_title)
