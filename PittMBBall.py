import datetime
import os
import openpyxl

today = datetime.date.today().strftime("%m/%d/%Y")

# retrieve games from excel s

eventFile = 'C:\\Users\\DanFang\\Desktop\\EventSeasons\\pittmbb.xlsx'
eventwb = openpyxl.load_workbook(eventFile, data_only = True, read_only = True)
sheet = eventwb.get_sheet_by_name('Sheet')
eventdates = []
eventtimes = []
opponents = []
for row in range(1, sheet.max_row + 1):
    eventdates.append(sheet['A' + str(row)].value.date())
    eventtimes.append((sheet['A' + str(row)].value + datetime.timedelta(seconds=1)).time().strftime('%I:%M %p').lstrip('0'))
    opponents.append(sheet['B' + str(row)].value[sheet['B' + str(row)].value.index('vs. ') + 4:])

print(eventdates)
print(eventtimes)
print(opponents)


downloadFolder = 'C:\\Users\\DanFang\\Downloads'
os.chdir(downloadFolder)

dated_files = [(os.path.getmtime(fn), os.path.basename(fn))
               for fn in os.listdir(downloadFolder) if fn.lower().endswith('.xlsx')]
dated_files.sort()
dated_files.reverse()
excel = dated_files[0][1]

workbook = openpyxl.load_workbook(excel, data_only = True, read_only = True)
sheet = workbook.get_sheet_by_name('Ticket Request Advanced Fin...')
print(sheet)


class Request(object):
    def __init__(self, event = "", agency = "", preferred_date = "", tickets_requested = "", chaperone_name = "", notes = "", status = "", created_on = ""):
        self.event = event
        self.agency = agency
        self.preferred_date = preferred_date
        self.tickets_requested = tickets_requested
        self.chaperone_name = chaperone_name
        self.notes = notes
        self.status = status
        self.created_on = created_on

request_objects = []
for row in range(2, sheet.max_row + 1):
    preferred_date = datetime.datetime.date(sheet['G' + str(row)].value)

    request_objects.append(Request(sheet['F' + str(row)].value, sheet['E' + str(row)].value, preferred_date, sheet['H' + str(row)].value, sheet['I' + str(row)].value, sheet['P' + str(row)].value, sheet['Q' + str(row)].value, sheet['D' + str(row)].value))

request_objects.sort(key=lambda r: r.created_on)
print(len(request_objects))
print(request_objects[0].preferred_date)

out_folder = "C:\\Users\\DanFang\\Desktop\\Requests"
os.chdir(out_folder)

#begin writing new spreadsheet

out_workbook = openpyxl.Workbook()

new_requests_style = openpyxl.styles.NamedStyle(name="new_request_style")
new_requests_style.fill = openpyxl.styles.PatternFill(patternType='solid', fill_type='solid',fgColor=openpyxl.styles.Color('afeeee'))
border = openpyxl.styles.Side(style='thin', color='000000')
new_requests_style.border = openpyxl.styles.Border(left=border, top=border, right=border, bottom=border)

out_workbook.add_named_style(new_requests_style)

old_requests_style = openpyxl.styles.NamedStyle(name="old_request_style")
border = openpyxl.styles.Side(style='thin', color='000000')
old_requests_style.border = openpyxl.styles.Border(left=border, top=border, right=border, bottom=border)

out_workbook.add_named_style(old_requests_style)


def populate_sheet(sheet,game_date,game_time,game_title,preferred_date,game_number):
    sheet['A1'] = 'Tickets for Kids'
    sheet['A2'] = 'From:'
    sheet['A3'] = 'Game Date:'
    sheet['A4'] = 'Game Time:'
    sheet['B2'] = 'Dan Fang'
    sheet['B2'].font = openpyxl.styles.Font(bold=True)
    sheet['B3'] = game_date
    sheet['B3'].font = openpyxl.styles.Font(bold=True)
    sheet['B4'] = game_time
    sheet['B4'].font = openpyxl.styles.Font(bold=True)
    sheet['D2'].style = new_requests_style
    sheet['D2'] = 'New Requests'

    sheet['A6'] = game_title
    sheet['A6'].font = openpyxl.styles.Font(bold=True)
    sheet['B6'] = 'Agency:'
    sheet['B6'].font = openpyxl.styles.Font(bold=True)
    sheet['C6'] = 'Number:'
    sheet['C6'].font = openpyxl.styles.Font(bold=True)
    sheet['D6'] = 'Will Call Name:'
    sheet['D6'].font = openpyxl.styles.Font(bold=True)
    sheet['E6'] = 'Special Notes:'
    sheet['E6'].font = openpyxl.styles.Font(bold=True)

    row_count = 7
    for request in request_objects:
        if str(request.preferred_date) == preferred_date:
            if request.status == 'Approved':
                # sheet["A" + str(row_count)] = game_number
                sheet["A" + str(row_count)].style = old_requests_style
                sheet["B" + str(row_count)] = request.agency
                sheet["B" + str(row_count)].style = old_requests_style
                sheet["C" + str(row_count)] = request.tickets_requested
                sheet["C" + str(row_count)].style = old_requests_style
                sheet["D" + str(row_count)] = request.chaperone_name
                sheet["D" + str(row_count)].style = old_requests_style
                sheet["E" + str(row_count)] = request.notes
                sheet["E" + str(row_count)].style = old_requests_style

                row_count += 1

    for request in request_objects:
        if str(request.preferred_date) == preferred_date:
            if request.status == 'Pending':
                # sheet["A" + str(row_count)] = game_number
                sheet["A" + str(row_count)].style = new_requests_style
                sheet["B" + str(row_count)] = request.agency
                sheet["B" + str(row_count)].style = new_requests_style
                sheet["C" + str(row_count)] = request.tickets_requested
                sheet["C" + str(row_count)].style = new_requests_style
                sheet["D" + str(row_count)] = request.chaperone_name
                sheet["D" + str(row_count)].style = new_requests_style
                sheet["E" + str(row_count)] = request.notes
                sheet["E" + str(row_count)].style = new_requests_style

                row_count += 1

    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 45
    sheet.column_dimensions["C"].width = 12
    sheet.column_dimensions["D"].width = 20
    sheet.column_dimensions["E"].width = 45


for i in range(0, len(eventdates)):
    sheet = out_workbook.create_sheet(title = eventdates[i].strftime('%D').replace('/', '-') + ' vs ' + opponents[i])
    populate_sheet(sheet, eventdates[i].strftime('%D'), eventtimes[i], 'Pitt vs ' + opponents[i], eventdates[i].strftime('%Y-%m-%d'), '')

out_workbook.remove_sheet(out_workbook.get_sheet_by_name('Sheet'))
out_workbook.save("TFK Pitt Men's Basketball 2017-2018 Requests.xlsx")
